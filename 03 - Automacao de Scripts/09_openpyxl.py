"""
======================================================
MÓDULO: openpyxl – Manipulação de Arquivos Excel (.xlsx)
======================================================

O `openpyxl` é a biblioteca Python mais usada para **ler, criar e modificar arquivos Excel do tipo .xlsx** (formato do Excel moderno).

É ideal para:
- Gerar relatórios e planilhas automaticamente
- Preencher modelos de Excel com dados de banco de dados ou APIs
- Ler e extrair dados de planilhas existentes
- Automatizar tarefas com arquivos do Excel

Instalação:
    pip install openpyxl

-----------------------------------------------
FUNCIONALIDADES PRINCIPAIS:
-----------------------------------------------
- Criar novas planilhas Excel
- Ler e escrever células
- Adicionar fórmulas, estilos, bordas, cores
- Trabalhar com múltiplas abas (worksheets)
- Ler arquivos existentes sem alterar o formato

-----------------------------------------------
EXEMPLO BÁSICO: CRIANDO UMA NOVA PLANILHA
-----------------------------------------------
"""

from openpyxl import Workbook

# Cria um novo arquivo Excel
wb = Workbook()

# Seleciona a planilha ativa
ws = wb.active
ws.title = "Relatório"

# Escreve valores em células
ws["A1"] = "Nome"
ws["B1"] = "Idade"
ws.append(["Ana", 28])
ws.append(["João", 32])

# Salva o arquivo
wb.save("relatorio.xlsx")

"""
-----------------------------------------------
EXEMPLO: LENDO UM ARQUIVO EXISTENTE
-----------------------------------------------
"""

from openpyxl import load_workbook

# Abre um arquivo existente
wb = load_workbook("relatorio.xlsx")
ws = wb["Relatório"]

# Lê os dados da planilha
for row in ws.iter_rows(values_only=True):
    print(row)

"""
-----------------------------------------------
DICAS:
-----------------------------------------------
- Use `ws.cell(row=1, column=2).value` para acessar células dinamicamente.
- Pode aplicar estilos com `openpyxl.styles` (cores, negrito, alinhamento etc.).
- É possível proteger planilhas, ocultar colunas, inserir gráficos e fórmulas.

Documentação oficial: https://openpyxl.readthedocs.io/
"""